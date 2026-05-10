import os
import tempfile
from datetime import datetime

import openpyxl
import xlrd


def read_excel_headers(file_path, header_row=0, sheet_name=None):
    """Read headers from a workbook and return headers plus sheet names."""
    ext = os.path.splitext(file_path)[1].lower()
    header_row = int(header_row)

    if ext == ".xlsx":
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheets = workbook.sheetnames
            worksheet = workbook[sheet_name] if sheet_name and sheet_name in sheets else workbook.active
            headers = [str(cell.value) if cell.value is not None else "" for cell in worksheet[header_row + 1]]
            return headers, sheets
        finally:
            workbook.close()

    if ext == ".xls":
        workbook = xlrd.open_workbook(file_path)
        sheets = workbook.sheet_names()
        worksheet = workbook.sheet_by_name(sheet_name) if sheet_name and sheet_name in sheets else workbook.sheet_by_index(0)
        headers = [str(worksheet.cell_value(header_row, col)) for col in range(worksheet.ncols)]
        return headers, sheets

    raise ValueError(f"Unsupported file format: {ext}")


def copy_excel_with_mapping(
    source_path,
    target_path,
    mapping,
    source_header_row,
    target_header_row,
    mode="replace",
    source_sheet=None,
    target_sheet=None,
    output_dir=None,
    output_name=None,
):
    """
    Copy source workbook data into target workbook according to column mapping.

    mapping uses {target_column_index: source_column_index}; both indexes are zero-based.
    """
    mapping = {int(target_col): int(source_col) for target_col, source_col in mapping.items()}
    source_header_row = int(source_header_row)
    target_header_row = int(target_header_row)

    source_ext = os.path.splitext(source_path)[1].lower()
    target_ext = os.path.splitext(target_path)[1].lower()
    source_data = read_source_rows(source_path, source_ext, source_sheet)

    if target_ext != ".xlsx":
        raise ValueError("Target file must be xlsx format.")

    target_workbook = openpyxl.load_workbook(target_path)
    try:
        target_worksheet = (
            target_workbook[target_sheet]
            if target_sheet and target_sheet in target_workbook.sheetnames
            else target_workbook.active
        )

        target_header_row_number = target_header_row + 1
        start_row = target_header_row_number + 1 if mode == "replace" else target_worksheet.max_row + 1

        ensure_writable_data_area(target_worksheet, start_row)

        if mode == "replace":
            clear_mapped_columns(target_worksheet, mapping, start_row)

        for offset, source_row_index in enumerate(range(source_header_row + 1, len(source_data))):
            target_row_number = target_header_row_number + 1 + offset if mode == "replace" else start_row + offset
            if target_row_number <= target_header_row_number:
                target_row_number = target_header_row_number + 1

            for target_col, source_col in mapping.items():
                source_row = source_data[source_row_index]
                value = source_row[source_col] if source_col < len(source_row) else None
                target_worksheet.cell(row=target_row_number, column=target_col + 1, value=value)

        output_root = output_dir or tempfile.gettempdir()
        os.makedirs(output_root, exist_ok=True)
        output_filename = output_name or f"mapped_{datetime.now():%Y%m%d_%H%M%S}_{os.getpid()}.xlsx"
        output_path = os.path.join(output_root, output_filename)
        target_workbook.save(output_path)
        return output_path
    finally:
        target_workbook.close()


def read_source_rows(source_path, source_ext, source_sheet=None):
    if source_ext == ".xlsx":
        workbook = openpyxl.load_workbook(source_path)
        try:
            worksheet = workbook[source_sheet] if source_sheet and source_sheet in workbook.sheetnames else workbook.active
            return [[cell.value for cell in row] for row in worksheet.rows]
        finally:
            workbook.close()

    if source_ext == ".xls":
        workbook = xlrd.open_workbook(source_path)
        worksheet = workbook.sheet_by_name(source_sheet) if source_sheet and source_sheet in workbook.sheet_names() else workbook.sheet_by_index(0)
        return [[worksheet.cell_value(row, col) for col in range(worksheet.ncols)] for row in range(worksheet.nrows)]

    raise ValueError(f"Unsupported file format: {source_ext}")


def ensure_writable_data_area(worksheet, start_row):
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.max_row >= start_row:
            raise Exception("Target data area contains merged cells. Unmerge them or use append mode.")


def clear_mapped_columns(worksheet, mapping, start_row):
    for row in worksheet.iter_rows(min_row=start_row, max_row=worksheet.max_row):
        for target_col in mapping:
            if target_col < len(row):
                row[target_col].value = None
